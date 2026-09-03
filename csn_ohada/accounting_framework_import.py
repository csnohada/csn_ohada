from __future__ import annotations

import csv
import io
import json
from collections import Counter
from pathlib import PurePosixPath

import frappe
from frappe import _


REQUIRED_COLUMNS = (
	"framework_code",
	"framework_version",
	"account_code",
	"account_label",
	"normal_balance",
	"valid_from",
	"legal_reference",
	"source_document",
	"status",
)

SUPPORTED_COLUMNS = (
	"framework_code",
	"framework_version",
	"account_code",
	"account_label",
	"parent_account_code",
	"account_class",
	"account_category",
	"normal_balance",
	"is_postable",
	"is_control_account",
	"requires_third_party",
	"requires_budget_line",
	"requires_project",
	"requires_fund",
	"requires_donor",
	"requires_campaign",
	"requires_emergency",
	"requires_cost_center",
	"requires_asset",
	"requires_inventory_item",
	"requires_bank_account",
	"valid_from",
	"valid_to",
	"legal_reference",
	"source_document",
	"status",
)

CHECK_COLUMNS = tuple(column for column in SUPPORTED_COLUMNS if column.startswith(("is_", "requires_")))
ALLOWED_ROLES = {"System Manager", "CSN Administrateur Referentiel"}


def _check_permission() -> None:
	roles = set(frappe.get_roles())
	if not roles.intersection(ALLOWED_ROLES):
		frappe.throw(_("Vous n'êtes pas autorisé à importer un référentiel comptable."), frappe.PermissionError)


def _get_file(file_url: str):
	file_name = frappe.db.get_value("File", {"file_url": file_url}, "name")
	if not file_name:
		frappe.throw(_("Fichier introuvable."))
	return frappe.get_doc("File", file_name)


def _parse_csv(content: bytes) -> list[dict]:
	text = content.decode("utf-8-sig")
	return list(csv.DictReader(io.StringIO(text)))


def _parse_json(content: bytes) -> list[dict]:
	payload = json.loads(content.decode("utf-8-sig"))
	if isinstance(payload, dict):
		payload = payload.get("accounts", payload.get("data"))
	if not isinstance(payload, list):
		frappe.throw(_("Le JSON doit contenir une liste de comptes ou une clé accounts/data."))
	return payload


def _parse_xlsx(content: bytes) -> list[dict]:
	from openpyxl import load_workbook

	workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
	sheet = workbook.active
	rows = sheet.iter_rows(values_only=True)
	try:
		headers = [str(value or "").strip() for value in next(rows)]
	except StopIteration:
		return []
	return [dict(zip(headers, values, strict=False)) for values in rows if any(value is not None for value in values)]


def _read_rows(file_url: str) -> list[dict]:
	file_doc = _get_file(file_url)
	content = file_doc.get_content()
	if isinstance(content, str):
		content = content.encode()
	extension = PurePosixPath(file_url.lower()).suffix
	if extension == ".csv":
		return _parse_csv(content)
	if extension == ".json":
		return _parse_json(content)
	if extension == ".xlsx":
		return _parse_xlsx(content)
	frappe.throw(_("Format non pris en charge. Utilisez XLSX, CSV ou JSON."))


def _as_check(value) -> int:
	return int(str(value or "0").strip().lower() in {"1", "true", "yes", "oui", "x"})


def _normalise(row: dict) -> dict:
	result = {key: row.get(key) for key in SUPPORTED_COLUMNS}
	for key, value in result.items():
		if isinstance(value, str):
			result[key] = value.strip()
	for key in CHECK_COLUMNS:
		result[key] = _as_check(result.get(key))
	result["account_code"] = str(result.get("account_code") or "").strip()
	result["parent_account_code"] = str(result.get("parent_account_code") or "").strip()
	return result


def _validate(rows: list[dict], framework_version: str) -> tuple[list[dict], list[dict]]:
	normalised = [_normalise(row) for row in rows]
	errors: list[dict] = []
	codes = [row["account_code"] for row in normalised if row["account_code"]]
	counts = Counter(codes)
	existing = set(
		frappe.get_all(
			"CSN Compte Referentiel",
			filters={"framework_version": framework_version},
			pluck="account_code",
		)
	)
	available = set(codes) | existing

	version = frappe.db.get_value(
		"CSN Version Referentiel Comptable", framework_version, ["framework", "status"], as_dict=True
	)
	if not version:
		frappe.throw(_("Version de référentiel introuvable."))
	if version.status in ("Active", "Inactive"):
		frappe.throw(_("Une version active ou historique ne peut plus recevoir d'import."))

	for index, row in enumerate(normalised, start=2):
		row_errors = []
		for column in REQUIRED_COLUMNS:
			if not row.get(column):
				row_errors.append(_("{0} est obligatoire").format(column))
		if row.get("framework_code") != version.framework:
			row_errors.append(_("framework_code ne correspond pas à la version sélectionnée"))
		if row.get("framework_version") != framework_version:
			row_errors.append(_("framework_version ne correspond pas à la version sélectionnée"))
		if counts[row["account_code"]] > 1:
			row_errors.append(_("code de compte dupliqué dans le fichier"))
		if row["account_code"] in existing:
			row_errors.append(_("code de compte déjà importé dans cette version"))
		parent = row.get("parent_account_code")
		if parent and parent not in available:
			row_errors.append(_("compte parent absent de la version et du fichier"))
		if parent and parent == row["account_code"]:
			row_errors.append(_("un compte ne peut pas être son propre parent"))
		if row.get("normal_balance") not in ("Débit", "Crédit"):
			row_errors.append(_("normal_balance doit être Débit ou Crédit"))
		if row_errors:
			errors.append({"row": index, "account_code": row["account_code"], "errors": row_errors})
	return normalised, errors


@frappe.whitelist()
def preview_account_import(file_url: str, framework_version: str) -> dict:
	_check_permission()
	rows, errors = _validate(_read_rows(file_url), framework_version)
	return {
		"valid": not errors,
		"total_rows": len(rows),
		"valid_rows": len(rows) - len(errors),
		"errors": errors,
		"preview": rows[:100],
	}


@frappe.whitelist(methods=["POST"])
def import_accounts(file_url: str, framework_version: str) -> dict:
	_check_permission()
	rows, errors = _validate(_read_rows(file_url), framework_version)
	if errors:
		frappe.throw(_("Import refusé : corrigez les erreurs signalées par la prévisualisation."))

	remaining = {row["account_code"]: row for row in rows}
	created: list[str] = []
	available = set(
		frappe.get_all(
			"CSN Compte Referentiel",
			filters={"framework_version": framework_version},
			pluck="account_code",
		)
	)

	while remaining:
		ready = [
			row for row in remaining.values() if not row.get("parent_account_code") or row["parent_account_code"] in available
		]
		if not ready:
			frappe.throw(_("Hiérarchie circulaire ou parent non résolu."))
		for row in ready:
			payload = {key: row.get(key) for key in SUPPORTED_COLUMNS if key not in ("framework_code", "framework_version")}
			payload.update({"doctype": "CSN Compte Referentiel", "framework_version": framework_version})
			document = frappe.get_doc(payload).insert()
			created.append(document.name)
			available.add(row["account_code"])
			remaining.pop(row["account_code"])

	frappe.db.set_value("CSN Version Referentiel Comptable", framework_version, "status", "Importée")
	return {"imported": len(created), "documents": created}
